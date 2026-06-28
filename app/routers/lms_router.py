from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form, BackgroundTasks
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import and_, or_, func, desc
from app import models, schemas, auth
from app.database import get_db
from typing import Optional, List
import json
import os
import uuid
import shutil
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

router = APIRouter(prefix="/api/lms", tags=["LMS"])

# Создаём папки для загрузок
UPLOAD_DIR = "app/static/uploads/lms"
ATTACHMENTS_DIR = os.path.join(UPLOAD_DIR, "attachments")
ANSWERS_DIR = os.path.join(UPLOAD_DIR, "answers")
LECTURES_DIR = os.path.join(UPLOAD_DIR, "lectures")

os.makedirs(ATTACHMENTS_DIR, exist_ok=True)
os.makedirs(ANSWERS_DIR, exist_ok=True)
os.makedirs(LECTURES_DIR, exist_ok=True)


# ========== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ==========

def get_file_type(filename: str) -> str:
    """Определяет тип файла по расширению"""
    ext = filename.split('.')[-1].lower() if '.' in filename else ''
    file_type_map = {
        'pdf': 'pdf',
        'doc': 'doc', 'docx': 'docx',
        'xls': 'xls', 'xlsx': 'xlsx',
        'zip': 'zip', 'rar': 'rar', '7z': 'archive',
        'mp4': 'video', 'avi': 'video', 'mov': 'video', 'mkv': 'video',
        'jpg': 'image', 'jpeg': 'image', 'png': 'image', 'webp': 'image', 'gif': 'image',
        'txt': 'text', 'csv': 'text',
        'ppt': 'ppt', 'pptx': 'ppt'
    }
    return file_type_map.get(ext, 'other')


def get_lecture_file_type(filename: str) -> str:
    """Определяет тип файла лекции"""
    ext = filename.split('.')[-1].lower() if '.' in filename else ''
    lecture_type_map = {
        'pdf': 'pdf',
        'doc': 'doc', 'docx': 'docx',
        'ppt': 'ppt', 'pptx': 'pptx',
        'txt': 'txt', 'md': 'md',
        'rtf': 'rtf'
    }
    return lecture_type_map.get(ext, 'other')


def update_course_progress(user_id: int, course_id: int, db: Session):
    """Обновляет общий прогресс курса на основе модулей и уроков"""
    modules = db.query(models.CourseModule).filter(
        models.CourseModule.course_id == course_id
    ).all()
    
    if not modules:
        return
    
    total_lessons = 0
    completed_lessons = 0
    
    for module in modules:
        lessons = db.query(models.CourseLesson).filter(
            models.CourseLesson.module_id == module.id
        ).all()
        
        for lesson in lessons:
            total_lessons += 1
            progress = db.query(models.UserLessonProgress).filter(
                and_(
                    models.UserLessonProgress.user_id == user_id,
                    models.UserLessonProgress.lesson_id == lesson.id,
                    models.UserLessonProgress.is_completed == True
                )
            ).first()
            if progress:
                completed_lessons += 1
    
    progress_percent = int((completed_lessons / total_lessons) * 100) if total_lessons > 0 else 0
    
    course_progress = db.query(models.UserProgress).filter(
        and_(
            models.UserProgress.user_id == user_id,
            models.UserProgress.course_id == course_id
        )
    ).first()
    
    if course_progress:
        course_progress.progress_percent = progress_percent
        if progress_percent == 100:
            course_progress.is_completed = True
            course_progress.completed_at = datetime.utcnow()
        course_progress.last_activity = datetime.utcnow()
    else:
        course_progress = models.UserProgress(
            user_id=user_id,
            course_id=course_id,
            progress_percent=progress_percent
        )
        db.add(course_progress)
    
    db.commit()


def reset_lesson_progress(user_id: int, lesson_id: int, db: Session):
    """Полностью сбрасывает прогресс ученика по уроку"""
    # Удаляем прогресс урока
    lesson_progress = db.query(models.UserLessonProgress).filter(
        and_(
            models.UserLessonProgress.user_id == user_id,
            models.UserLessonProgress.lesson_id == lesson_id
        )
    ).first()
    if lesson_progress:
        db.delete(lesson_progress)
    
    # Находим задание к уроку
    lesson = db.query(models.CourseLesson).filter(models.CourseLesson.id == lesson_id).first()
    if lesson and lesson.assignment:
        # Удаляем все ответы
        submissions = db.query(models.AssignmentSubmission).filter(
            and_(
                models.AssignmentSubmission.assignment_id == lesson.assignment.id,
                models.AssignmentSubmission.user_id == user_id
            )
        ).all()
        
        for submission in submissions:
            # Удаляем ответы на вопросы
            db.query(models.UserAnswer).filter(
                models.UserAnswer.submission_id == submission.id
            ).delete()
            # Удаляем попытки
            db.query(models.AssignmentAttempt).filter(
                models.AssignmentAttempt.submission_id == submission.id
            ).delete()
            # Удаляем саму submission
            db.delete(submission)
    
    db.commit()
    
    # Обновляем прогресс курса
    if lesson:
        update_course_progress(user_id, lesson.module.course_id, db)


def get_remaining_attempts(assignment_id: int, user_id: int, db: Session) -> int:
    """Получить количество оставшихся попыток"""
    assignment = db.query(models.LessonAssignment).filter(
        models.LessonAssignment.id == assignment_id
    ).first()
    if not assignment:
        return 0
    
    if assignment.max_attempts == 0:
        return 999  # Без ограничений
    
    attempts_count = db.query(models.AssignmentAttempt).filter(
        and_(
            models.AssignmentAttempt.submission_id.in_(
                db.query(models.AssignmentSubmission.id).filter(
                    models.AssignmentSubmission.assignment_id == assignment_id,
                    models.AssignmentSubmission.user_id == user_id
                )
            )
        )
    ).count()
    
    return max(0, assignment.max_attempts - attempts_count)


def get_assignment_status_internal(assignment_id: int, user_id: int, db: Session):
    """Внутренняя функция для получения статуса задания"""
    assignment = db.query(models.LessonAssignment).filter(
        models.LessonAssignment.id == assignment_id
    ).first()
    if not assignment:
        return None
    
    submission = db.query(models.AssignmentSubmission).filter(
        and_(
            models.AssignmentSubmission.assignment_id == assignment_id,
            models.AssignmentSubmission.user_id == user_id,
            models.AssignmentSubmission.is_latest == True
        )
    ).first()
    
    if not submission:
        return {
            "status": "not_started",
            "message": "Assignment not started",
            "score": None,
            "max_score": assignment.max_score,
            "passing_score": assignment.passing_score,
            "is_passed": False,
            "attempt": 0,
            "max_attempts": assignment.max_attempts,
            "submitted_at": None,
            "graded_at": None,
            "teacher_comment": None,
            "questions": [],
            "allow_retake": assignment.allow_retake,
            "retake_count": assignment.retake_count,
            "deadline": assignment.deadline,
            "is_overdue": False,
            "is_retake": False,
            "retake_number": 0,
            "can_retake": True,
            "remaining_attempts": assignment.max_attempts if assignment.max_attempts > 0 else 999
        }
    
    questions = db.query(models.AssignmentQuestion).filter(
        models.AssignmentQuestion.assignment_id == assignment_id
    ).all()
    
    question_status = []
    all_answered = True
    has_pending = False
    
    for question in questions:
        answer = db.query(models.UserAnswer).filter(
            and_(
                models.UserAnswer.submission_id == submission.id,
                models.UserAnswer.question_id == question.id
            )
        ).first()
        
        status_item = {
            "question_id": question.id,
            "question_type": question.question_type,
            "is_answered": answer is not None,
            "is_correct": answer.is_correct if answer else None,
            "points_earned": answer.points_earned if answer else 0,
            "answer_text": answer.answer_text if answer else None,
            "answer_file": answer.answer_file if answer else None
        }
        
        if answer and answer.is_correct is None and (answer.answer_text or answer.answer_file):
            status_item["status"] = "pending"
            has_pending = True
        elif answer and answer.is_correct is not None:
            status_item["status"] = "graded"
        elif not answer:
            status_item["status"] = "not_answered"
            all_answered = False
        else:
            status_item["status"] = "unknown"
        
        question_status.append(status_item)
    
    is_overdue = False
    if assignment.deadline and assignment.deadline < datetime.utcnow():
        is_overdue = True
    
    # Проверяем, можно ли пересдать
    can_retake = assignment.allow_retake
    remaining_attempts = get_remaining_attempts(assignment_id, user_id, db)
    if assignment.max_attempts > 0 and remaining_attempts <= 0:
        can_retake = False
    
    if submission.is_passed:
        overall_status = "passed"
        message = f"✅ Задание пройдено! Баллы: {submission.score}/{assignment.max_score}"
    elif has_pending:
        overall_status = "pending_teacher"
        message = "⏳ Ожидает проверки учителем"
    elif submission.score is not None and submission.score < assignment.passing_score:
        if submission.teacher_comment and "пересдачу" in submission.teacher_comment.lower():
            overall_status = "retake"
            message = f"🔄 Отправлено на пересдачу. Комментарий: {submission.teacher_comment}"
        elif can_retake and remaining_attempts > 0:
            overall_status = "failed"
            message = f"❌ Не пройдено. Баллы: {submission.score}/{assignment.passing_score}. Осталось попыток: {remaining_attempts}"
        else:
            overall_status = "failed"
            message = f"❌ Не пройдено. Баллы: {submission.score}/{assignment.passing_score}. Попыток больше нет."
    elif not all_answered:
        overall_status = "in_progress"
        message = "✏️ Ответьте на все вопросы"
    else:
        overall_status = "submitted"
        message = "📤 Отправлено на проверку"
    
    return {
        "status": overall_status,
        "message": message,
        "score": submission.score,
        "max_score": assignment.max_score,
        "passing_score": assignment.passing_score,
        "is_passed": submission.is_passed,
        "attempt": submission.current_attempt,
        "max_attempts": assignment.max_attempts,
        "submitted_at": submission.submitted_at,
        "graded_at": submission.graded_at,
        "teacher_comment": submission.teacher_comment,
        "questions": question_status,
        "allow_retake": assignment.allow_retake,
        "retake_count": assignment.retake_count,
        "deadline": assignment.deadline,
        "is_overdue": is_overdue,
        "is_retake": submission.is_retake,
        "retake_number": submission.retake_number,
        "can_retake": can_retake,
        "remaining_attempts": remaining_attempts
    }


# ========== ПОЛУЧЕНИЕ ДАННЫХ КУРСА ==========

@router.get("/courses/{course_id}/full")
def get_course_full(
    course_id: int,
    db: Session = Depends(get_db),
    current_user: Optional[models.User] = Depends(auth.get_current_user_optional)
):
    """Получить полную информацию о курсе с модулями и уроками"""
    course = db.query(models.Course).filter(models.Course.id == course_id).first()
    if not course:
        raise HTTPException(status_code=404, detail="Course not found")
    
    modules = db.query(models.CourseModule).filter(
        models.CourseModule.course_id == course_id
    ).order_by(models.CourseModule.order_index).all()
    
    result = {
        "id": course.id,
        "title": course.title,
        "description": course.description,
        "image_url": course.image_url,
        "price": course.price,
        "format_type": course.format_type,
        "modules": []
    }
    
    for module in modules:
        lessons = db.query(models.CourseLesson).filter(
            models.CourseLesson.module_id == module.id
        ).order_by(models.CourseLesson.order_index).all()
        
        module_data = {
            "id": module.id,
            "title": module.title,
            "description": module.description,
            "module_type": module.module_type,
            "order_index": module.order_index,
            "lessons": []
        }
        
        for lesson in lessons:
            is_completed = False
            progress_data = None
            has_pending_submission = False
            
            if current_user:
                progress = db.query(models.UserLessonProgress).filter(
                    and_(
                        models.UserLessonProgress.user_id == current_user.id,
                        models.UserLessonProgress.lesson_id == lesson.id
                    )
                ).first()
                if progress:
                    is_completed = progress.is_completed
                    progress_data = {
                        "last_position": progress.last_position,
                        "video_watched_percent": progress.video_watched_percent,
                        "lecture_file_downloaded": progress.lecture_file_downloaded
                    }
                
                # Проверяем наличие непроверенных ответов
                if lesson.assignment:
                    pending = db.query(models.AssignmentSubmission).filter(
                        and_(
                            models.AssignmentSubmission.assignment_id == lesson.assignment.id,
                            models.AssignmentSubmission.user_id == current_user.id,
                            models.AssignmentSubmission.is_passed == False,
                            models.AssignmentSubmission.graded_at.is_(None)
                        )
                    ).first()
                    has_pending_submission = pending is not None
            
            assignment = db.query(models.LessonAssignment).filter(
                models.LessonAssignment.lesson_id == lesson.id
            ).first()
            
            # Проверяем дедлайн
            is_overdue = False
            if lesson.deadline and lesson.deadline < datetime.utcnow():
                is_overdue = True
            
            module_data["lessons"].append({
                "id": lesson.id,
                "title": lesson.title,
                "content": lesson.content,
                "video_url": lesson.video_url,
                "is_free": lesson.is_free,
                "duration_minutes": lesson.duration_minutes,
                "has_assignment": assignment is not None,
                "is_completed": is_completed,
                "has_pending_submission": has_pending_submission,
                "is_overdue": is_overdue,
                "deadline": lesson.deadline,
                "progress": progress_data,
                "lecture_type": lesson.lecture_type,
                "has_lecture_file": lesson.has_lecture_file,
                "lecture_file_name": lesson.lecture_file_name,
                "lecture_file_type": lesson.lecture_file_type,
                "allow_retake": lesson.allow_retake,
                "attachments": [{
                    "id": a.id,
                    "filename": a.filename,
                    "file_url": a.file_url,
                    "file_type": a.file_type,
                    "is_required": a.is_required,
                    "file_size": a.file_size
                } for a in lesson.attachments]
            })
        
        result["modules"].append(module_data)
    
    return result


@router.get("/courses/{course_id}/progress")
def get_course_progress(
    course_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Получить детальный прогресс по курсу"""
    course = db.query(models.Course).filter(models.Course.id == course_id).first()
    if not course:
        raise HTTPException(status_code=404, detail="Course not found")
    
    modules = db.query(models.CourseModule).filter(
        models.CourseModule.course_id == course_id
    ).order_by(models.CourseModule.order_index).all()
    
    result = {
        "course_id": course.id,
        "course_title": course.title,
        "total_lessons": 0,
        "completed_lessons": 0,
        "progress_percent": 0,
        "modules": []
    }
    
    total_lessons = 0
    completed_lessons = 0
    
    for module in modules:
        lessons = db.query(models.CourseLesson).filter(
            models.CourseLesson.module_id == module.id
        ).all()
        
        module_completed = 0
        module_total = len(lessons)
        
        for lesson in lessons:
            total_lessons += 1
            progress = db.query(models.UserLessonProgress).filter(
                and_(
                    models.UserLessonProgress.user_id == current_user.id,
                    models.UserLessonProgress.lesson_id == lesson.id,
                    models.UserLessonProgress.is_completed == True
                )
            ).first()
            if progress:
                completed_lessons += 1
                module_completed += 1
        
        result["modules"].append({
            "module_id": module.id,
            "module_title": module.title,
            "total_lessons": module_total,
            "completed_lessons": module_completed,
            "progress_percent": int((module_completed / module_total) * 100) if module_total > 0 else 0
        })
    
    result["total_lessons"] = total_lessons
    result["completed_lessons"] = completed_lessons
    result["progress_percent"] = int((completed_lessons / total_lessons) * 100) if total_lessons > 0 else 0
    
    return result


# ========== УПРАВЛЕНИЕ МОДУЛЯМИ ==========

@router.post("/courses/{course_id}/modules")
def create_module(
    course_id: int,
    module_data: schemas.ModuleCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_admin)
):
    """Создать модуль курса"""
    course = db.query(models.Course).filter(models.Course.id == course_id).first()
    if not course:
        raise HTTPException(status_code=404, detail="Course not found")
    
    module = models.CourseModule(
        course_id=course_id,
        title=module_data.title,
        description=module_data.description,
        module_type=module_data.module_type,
        order_index=module_data.order_index
    )
    db.add(module)
    db.commit()
    db.refresh(module)
    
    return {"message": "Module created", "id": module.id, "title": module.title}


@router.put("/modules/{module_id}")
def update_module(
    module_id: int,
    module_data: schemas.ModuleUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_admin)
):
    """Обновить модуль"""
    module = db.query(models.CourseModule).filter(models.CourseModule.id == module_id).first()
    if not module:
        raise HTTPException(status_code=404, detail="Module not found")
    
    for key, value in module_data.model_dump(exclude_unset=True).items():
        setattr(module, key, value)
    
    db.commit()
    return {"message": "Module updated"}


@router.delete("/modules/{module_id}")
def delete_module(
    module_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_admin)
):
    """Удалить модуль"""
    module = db.query(models.CourseModule).filter(models.CourseModule.id == module_id).first()
    if not module:
        raise HTTPException(status_code=404, detail="Module not found")
    
    db.delete(module)
    db.commit()
    return {"message": "Module deleted"}


# ========== УПРАВЛЕНИЕ УРОКАМИ ==========

@router.post("/modules/{module_id}/lessons")
def create_lesson(
    module_id: int,
    lesson_data: schemas.LessonCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_admin)
):
    """Создать урок в модуле"""
    module = db.query(models.CourseModule).filter(models.CourseModule.id == module_id).first()
    if not module:
        raise HTTPException(status_code=404, detail="Module not found")
    
    lesson = models.CourseLesson(
        module_id=module_id,
        title=lesson_data.title,
        content=lesson_data.content,
        video_url=lesson_data.video_url,
        is_free=lesson_data.is_free,
        duration_minutes=lesson_data.duration_minutes,
        order_index=lesson_data.order_index,
        lecture_type=lesson_data.lecture_type,
        allow_retake=lesson_data.allow_retake,
        deadline=lesson_data.deadline,
        is_published=lesson_data.is_published
    )
    db.add(lesson)
    db.commit()
    db.refresh(lesson)
    
    return {"message": "Lesson created", "id": lesson.id, "title": lesson.title}


@router.put("/lessons/{lesson_id}")
def update_lesson(
    lesson_id: int,
    lesson_data: schemas.LessonUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_admin)
):
    """Обновить урок"""
    lesson = db.query(models.CourseLesson).filter(models.CourseLesson.id == lesson_id).first()
    if not lesson:
        raise HTTPException(status_code=404, detail="Lesson not found")
    
    for key, value in lesson_data.model_dump(exclude_unset=True).items():
        setattr(lesson, key, value)
    
    db.commit()
    return {"message": "Lesson updated"}


@router.delete("/lessons/{lesson_id}")
def delete_lesson(
    lesson_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_admin)
):
    """Удалить урок"""
    lesson = db.query(models.CourseLesson).filter(models.CourseLesson.id == lesson_id).first()
    if not lesson:
        raise HTTPException(status_code=404, detail="Lesson not found")
    
    db.delete(lesson)
    db.commit()
    return {"message": "Lesson deleted"}


# ========== ЗАВЕРШЕНИЕ УРОКА ==========

@router.post("/lessons/{lesson_id}/complete")
def complete_lesson(
    lesson_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Отметить урок как пройденный"""
    lesson = db.query(models.CourseLesson).filter(models.CourseLesson.id == lesson_id).first()
    if not lesson:
        raise HTTPException(status_code=404, detail="Lesson not found")
    
    # Проверяем, есть ли задание и пройдено ли оно
    if lesson.assignment:
        submission = db.query(models.AssignmentSubmission).filter(
            and_(
                models.AssignmentSubmission.assignment_id == lesson.assignment.id,
                models.AssignmentSubmission.user_id == current_user.id,
                models.AssignmentSubmission.is_latest == True,
                models.AssignmentSubmission.is_passed == True
            )
        ).first()
        
        if not submission:
            raise HTTPException(status_code=403, detail="Assignment not passed yet")
    
    # Проверяем, не пройден ли уже урок
    existing_progress = db.query(models.UserLessonProgress).filter(
        and_(
            models.UserLessonProgress.user_id == current_user.id,
            models.UserLessonProgress.lesson_id == lesson_id
        )
    ).first()
    
    if existing_progress and existing_progress.is_completed:
        return {"message": "Lesson already completed", "is_completed": True}
    
    # Создаем или обновляем прогресс
    if existing_progress:
        existing_progress.is_completed = True
        existing_progress.completed_at = datetime.utcnow()
    else:
        progress = models.UserLessonProgress(
            user_id=current_user.id,
            lesson_id=lesson_id,
            is_completed=True,
            completed_at=datetime.utcnow()
        )
        db.add(progress)
    
    db.commit()
    
    # Обновляем прогресс курса
    update_course_progress(current_user.id, lesson.module.course_id, db)
    
    return {"message": "Lesson completed successfully", "is_completed": True}


# ========== ПОЛУЧЕНИЕ УРОКА (ОСНОВНОЙ ЭНДПОИНТ) ==========

@router.get("/lessons/{lesson_id}")
def get_lesson(
    lesson_id: int,
    db: Session = Depends(get_db),
    current_user: Optional[models.User] = Depends(auth.get_current_user_optional)
):
    """Получить урок с содержимым и заданием"""
    lesson = db.query(models.CourseLesson).filter(models.CourseLesson.id == lesson_id).first()
    if not lesson:
        raise HTTPException(status_code=404, detail="Lesson not found")
    
    if not lesson.is_free and not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")
    
    if current_user and not lesson.is_free:
        registration = db.query(models.CourseRegistration).filter(
            and_(
                models.CourseRegistration.user_id == current_user.id,
                models.CourseRegistration.course_id == lesson.module.course_id
            )
        ).first()
        if not registration:
            raise HTTPException(status_code=403, detail="You are not registered for this course")
    
    is_completed = False
    progress_data = None
    
    if current_user:
        progress = db.query(models.UserLessonProgress).filter(
            and_(
                models.UserLessonProgress.user_id == current_user.id,
                models.UserLessonProgress.lesson_id == lesson_id
            )
        ).first()
        if progress:
            is_completed = progress.is_completed
            progress_data = {
                "last_position": progress.last_position,
                "video_watched_percent": progress.video_watched_percent,
                "completed_at": progress.completed_at,
                "lecture_file_downloaded": progress.lecture_file_downloaded,
                "time_spent_seconds": progress.time_spent_seconds
            }
    
    # Получаем задание
    assignment = db.query(models.LessonAssignment).filter(
        models.LessonAssignment.lesson_id == lesson_id
    ).first()
    
    assignment_data = None
    if assignment:
        questions = db.query(models.AssignmentQuestion).filter(
            models.AssignmentQuestion.assignment_id == assignment.id
        ).order_by(models.AssignmentQuestion.order_index).all()
        
        assignment_data = {
            "id": assignment.id,
            "title": assignment.title,
            "description": assignment.description,
            "assignment_type": assignment.assignment_type,
            "max_score": assignment.max_score,
            "passing_score": assignment.passing_score,
            "time_limit_minutes": assignment.time_limit_minutes,
            "show_timer": assignment.show_timer,
            "max_attempts": assignment.max_attempts,
            "deadline": assignment.deadline,
            "allow_retake": assignment.allow_retake,
            "retake_count": assignment.retake_count,
            "auto_grade": assignment.auto_grade,
            "questions": []
        }
        
        for q in questions:
            options_value = q.options
            if options_value and isinstance(options_value, str):
                try:
                    parsed = json.loads(options_value)
                    if isinstance(parsed, list):
                        options_value = parsed
                    elif isinstance(parsed, str):
                        try:
                            double_parsed = json.loads(parsed)
                            if isinstance(double_parsed, list):
                                options_value = double_parsed
                            else:
                                options_value = parsed
                        except:
                            options_value = parsed
                except:
                    pass
            
            assignment_data["questions"].append({
                "id": q.id,
                "question_text": q.question_text,
                "question_image": q.question_image,
                "question_video": q.question_video,
                "question_type": q.question_type,
                "options": options_value,
                "points": q.points,
                "is_required": q.is_required,
                "order_index": q.order_index,
                "hint": q.hint,
                "explanation": q.explanation
            })
    
    # Предыдущий и следующий урок
    prev_lesson = db.query(models.CourseLesson).filter(
        models.CourseLesson.module_id == lesson.module_id,
        models.CourseLesson.order_index < lesson.order_index
    ).order_by(models.CourseLesson.order_index.desc()).first()
    
    next_lesson = db.query(models.CourseLesson).filter(
        models.CourseLesson.module_id == lesson.module_id,
        models.CourseLesson.order_index > lesson.order_index
    ).order_by(models.CourseLesson.order_index.asc()).first()
    
    # Получаем статус задания для текущего пользователя
    assignment_status = None
    if current_user and assignment:
        try:
            assignment_status = get_assignment_status_internal(assignment.id, current_user.id, db)
        except Exception as e:
            print(f"Error getting assignment status: {e}")
    
    return {
        "id": lesson.id,
        "title": lesson.title,
        "content": lesson.content,
        "video_url": lesson.video_url,
        "duration_minutes": lesson.duration_minutes,
        "is_free": lesson.is_free,
        "is_completed": is_completed,
        "progress": progress_data,
        "has_assignment": assignment is not None,
        "attachments": [{
            "id": a.id,
            "filename": a.filename,
            "file_url": a.file_url,
            "file_type": a.file_type,
            "is_required": a.is_required,
            "file_size": a.file_size
        } for a in lesson.attachments],
        "assignment": assignment_data,
        "prev_lesson_id": prev_lesson.id if prev_lesson else None,
        "next_lesson_id": next_lesson.id if next_lesson else None,
        "lecture_type": lesson.lecture_type,
        "lecture_file_url": lesson.lecture_file_url,
        "lecture_file_name": lesson.lecture_file_name,
        "lecture_file_type": lesson.lecture_file_type,
        "lecture_file_size": lesson.lecture_file_size,
        "has_lecture_file": lesson.has_lecture_file,
        "allow_retake": lesson.allow_retake,
        "deadline": lesson.deadline,
        "is_published": lesson.is_published,
        "assignment_status": assignment_status
    }


# ========== УПРАВЛЕНИЕ ФАЙЛАМИ ЛЕКЦИЙ ==========

@router.post("/lessons/{lesson_id}/upload-lecture")
async def upload_lecture_file(
    lesson_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_admin)
):
    """Загрузить файл лекции для урока"""
    lesson = db.query(models.CourseLesson).filter(models.CourseLesson.id == lesson_id).first()
    if not lesson:
        raise HTTPException(status_code=404, detail="Lesson not found")
    
    content = await file.read()
    file_size = len(content)
    
    if file_size > 50 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large (max 50MB)")
    
    file_type = get_lecture_file_type(file.filename)
    if file_type == 'other':
        raise HTTPException(status_code=400, detail="Unsupported file type. Use PDF, DOC, DOCX, PPT, PPTX, TXT or MD")
    
    ext = file.filename.split('.')[-1].lower()
    filename = f"{uuid.uuid4().hex}.{ext}"
    filepath = os.path.join(LECTURES_DIR, filename)
    
    with open(filepath, "wb") as buffer:
        buffer.write(content)
    
    file_url = f"/static/uploads/lms/lectures/{filename}"
    
    if lesson.lecture_file_url:
        old_file = lesson.lecture_file_url.replace("/static/uploads/lms/lectures/", "")
        old_path = os.path.join(LECTURES_DIR, old_file)
        if os.path.exists(old_path):
            os.remove(old_path)
    
    lesson.lecture_file_url = file_url
    lesson.lecture_file_name = file.filename
    lesson.lecture_file_type = file_type
    lesson.lecture_file_size = file_size
    lesson.has_lecture_file = True
    lesson.lecture_type = models.LessonLectureType.FILE
    
    db.commit()
    
    return {
        "message": "Lecture file uploaded",
        "file_url": file_url,
        "file_name": file.filename,
        "file_type": file_type,
        "file_size": file_size
    }


@router.delete("/lessons/{lesson_id}/remove-lecture")
def remove_lecture_file(
    lesson_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_admin)
):
    """Удалить файл лекции"""
    lesson = db.query(models.CourseLesson).filter(models.CourseLesson.id == lesson_id).first()
    if not lesson:
        raise HTTPException(status_code=404, detail="Lesson not found")
    
    if lesson.lecture_file_url:
        filepath = lesson.lecture_file_url.replace("/static/uploads/lms/lectures/", "")
        full_path = os.path.join(LECTURES_DIR, filepath)
        if os.path.exists(full_path):
            os.remove(full_path)
    
    lesson.lecture_file_url = None
    lesson.lecture_file_name = None
    lesson.lecture_file_type = None
    lesson.lecture_file_size = 0
    lesson.has_lecture_file = False
    
    db.commit()
    
    return {"message": "Lecture file removed"}


# ========== ЗАГРУЗКА ФАЙЛА ЛЕКЦИИ ДЛЯ УЧЕНИКА ==========

@router.get("/lessons/{lesson_id}/download-lecture")
def download_lecture_file(
    lesson_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Скачать файл лекции (для ученика)"""
    lesson = db.query(models.CourseLesson).filter(models.CourseLesson.id == lesson_id).first()
    if not lesson:
        raise HTTPException(status_code=404, detail="Lesson not found")
    
    if not lesson.has_lecture_file or not lesson.lecture_file_url:
        raise HTTPException(status_code=404, detail="Lecture file not found")
    
    if not lesson.is_free:
        registration = db.query(models.CourseRegistration).filter(
            and_(
                models.CourseRegistration.user_id == current_user.id,
                models.CourseRegistration.course_id == lesson.module.course_id
            )
        ).first()
        if not registration:
            raise HTTPException(status_code=403, detail="You are not registered for this course")
    
    progress = db.query(models.UserLessonProgress).filter(
        and_(
            models.UserLessonProgress.user_id == current_user.id,
            models.UserLessonProgress.lesson_id == lesson_id
        )
    ).first()
    
    if progress:
        progress.lecture_file_downloaded = True
        db.commit()
    
    filepath = lesson.lecture_file_url.replace("/static/uploads/lms/lectures/", "")
    full_path = os.path.join(LECTURES_DIR, filepath)
    
    if not os.path.exists(full_path):
        raise HTTPException(status_code=404, detail="File not found on server")
    
    return StreamingResponse(
        open(full_path, "rb"),
        media_type="application/octet-stream",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{lesson.lecture_file_name}"
        }
    )


# ========== СБРОС ПРОГРЕССА УРОКА (ДЛЯ УЧЕНИКА) ==========

@router.post("/lessons/{lesson_id}/reset-progress")
def reset_lesson_progress_endpoint(
    lesson_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Сбросить прогресс урока и пройти заново (для ученика)"""
    lesson = db.query(models.CourseLesson).filter(models.CourseLesson.id == lesson_id).first()
    if not lesson:
        raise HTTPException(status_code=404, detail="Lesson not found")
    
    if not lesson.allow_retake:
        raise HTTPException(status_code=403, detail="Retake is not allowed for this lesson")
    
    registration = db.query(models.CourseRegistration).filter(
        and_(
            models.CourseRegistration.user_id == current_user.id,
            models.CourseRegistration.course_id == lesson.module.course_id
        )
    ).first()
    if not registration:
        raise HTTPException(status_code=403, detail="You are not registered for this course")
    
    reset_lesson_progress(current_user.id, lesson_id, db)
    
    return {
        "message": "Progress reset successfully. You can retake the lesson.",
        "lesson_id": lesson_id,
        "allow_retake": lesson.allow_retake
    }


# ========== УПРАВЛЕНИЕ ВЛОЖЕНИЯМИ ==========

@router.post("/lessons/{lesson_id}/attachments")
async def add_lesson_attachment(
    lesson_id: int,
    file: UploadFile = File(...),
    is_required: bool = Form(False),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_admin)
):
    """Добавить прикрепление к уроку"""
    lesson = db.query(models.CourseLesson).filter(models.CourseLesson.id == lesson_id).first()
    if not lesson:
        raise HTTPException(status_code=404, detail="Lesson not found")
    
    content = await file.read()
    file_size = len(content)
    
    if file_size > 100 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large (max 100MB)")
    
    file_type = get_file_type(file.filename)
    
    ext = file.filename.split('.')[-1].lower()
    filename = f"{uuid.uuid4().hex}.{ext}"
    filepath = os.path.join(ATTACHMENTS_DIR, filename)
    
    with open(filepath, "wb") as buffer:
        buffer.write(content)
    
    file_url = f"/static/uploads/lms/attachments/{filename}"
    
    attachment = models.LessonAttachment(
        lesson_id=lesson_id,
        filename=file.filename,
        file_url=file_url,
        file_size=file_size,
        file_type=file_type,
        is_required=is_required
    )
    db.add(attachment)
    db.commit()
    db.refresh(attachment)
    
    return {
        "message": "Attachment added",
        "id": attachment.id,
        "file_url": file_url,
        "filename": file.filename
    }


@router.delete("/attachments/{attachment_id}")
def delete_attachment(
    attachment_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_admin)
):
    """Удалить прикрепление"""
    attachment = db.query(models.LessonAttachment).filter(
        models.LessonAttachment.id == attachment_id
    ).first()
    if not attachment:
        raise HTTPException(status_code=404, detail="Attachment not found")
    
    filepath = attachment.file_url.replace("/static/uploads/lms/attachments/", "")
    full_path = os.path.join(ATTACHMENTS_DIR, filepath)
    if os.path.exists(full_path):
        os.remove(full_path)
    
    db.delete(attachment)
    db.commit()
    return {"message": "Attachment deleted"}


# ========== УПРАВЛЕНИЕ ЗАДАНИЯМИ ==========

@router.post("/lessons/{lesson_id}/assignment")
def create_assignment(
    lesson_id: int,
    assignment_data: schemas.LessonAssignmentCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_admin)
):
    """Создать задание для урока с вопросами"""
    lesson = db.query(models.CourseLesson).filter(models.CourseLesson.id == lesson_id).first()
    if not lesson:
        raise HTTPException(status_code=404, detail="Lesson not found")
    
    existing = db.query(models.LessonAssignment).filter(
        models.LessonAssignment.lesson_id == lesson_id
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="Lesson already has an assignment")
    
    assignment = models.LessonAssignment(
        lesson_id=lesson_id,
        title=assignment_data.title,
        description=assignment_data.description,
        assignment_type=assignment_data.assignment_type,
        max_score=assignment_data.max_score,
        passing_score=assignment_data.passing_score,
        time_limit_minutes=assignment_data.time_limit_minutes,
        show_timer=assignment_data.show_timer,
        max_attempts=assignment_data.max_attempts,
        deadline=assignment_data.deadline,
        allow_retake=assignment_data.allow_retake,
        retake_count=assignment_data.retake_count,
        auto_grade=assignment_data.auto_grade
    )
    db.add(assignment)
    db.commit()
    db.refresh(assignment)
    
    for q_data in assignment_data.questions:
        options_value = json.dumps(q_data.options) if q_data.options else None
        
        question = models.AssignmentQuestion(
            assignment_id=assignment.id,
            question_text=q_data.question_text,
            question_image=q_data.question_image,
            question_video=q_data.question_video,
            question_type=q_data.question_type,
            options=options_value,
            correct_answer=q_data.correct_answer,
            points=q_data.points,
            order_index=q_data.order_index,
            is_required=q_data.is_required,
            hint=q_data.hint,
            explanation=q_data.explanation
        )
        db.add(question)
    
    db.commit()
    
    return {
        "message": "Assignment created",
        "id": assignment.id,
        "questions_count": len(assignment_data.questions)
    }


@router.put("/assignments/{assignment_id}")
def update_assignment(
    assignment_id: int,
    assignment_data: schemas.LessonAssignmentUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_admin)
):
    """Обновить задание"""
    assignment = db.query(models.LessonAssignment).filter(
        models.LessonAssignment.id == assignment_id
    ).first()
    if not assignment:
        raise HTTPException(status_code=404, detail="Assignment not found")
    
    for key, value in assignment_data.model_dump(exclude_unset=True).items():
        setattr(assignment, key, value)
    
    db.commit()
    return {"message": "Assignment updated"}


@router.delete("/assignments/{assignment_id}")
def delete_assignment(
    assignment_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_admin)
):
    """Удалить задание"""
    assignment = db.query(models.LessonAssignment).filter(
        models.LessonAssignment.id == assignment_id
    ).first()
    if not assignment:
        raise HTTPException(status_code=404, detail="Assignment not found")
    
    db.delete(assignment)
    db.commit()
    return {"message": "Assignment deleted"}


# ========== УПРАВЛЕНИЕ ВОПРОСАМИ ==========

@router.post("/assignments/{assignment_id}/questions")
def add_question(
    assignment_id: int,
    question_data: schemas.AssignmentQuestionCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_admin)
):
    """Добавить вопрос в задание"""
    assignment = db.query(models.LessonAssignment).filter(
        models.LessonAssignment.id == assignment_id
    ).first()
    if not assignment:
        raise HTTPException(status_code=404, detail="Assignment not found")
    
    options_value = json.dumps(question_data.options) if question_data.options else None
    
    question = models.AssignmentQuestion(
        assignment_id=assignment_id,
        question_text=question_data.question_text,
        question_image=question_data.question_image,
        question_video=question_data.question_video,
        question_type=question_data.question_type,
        options=options_value,
        correct_answer=question_data.correct_answer,
        points=question_data.points,
        order_index=question_data.order_index,
        is_required=question_data.is_required,
        hint=question_data.hint,
        explanation=question_data.explanation
    )
    db.add(question)
    db.commit()
    db.refresh(question)
    
    return {"message": "Question added", "id": question.id}


@router.put("/questions/{question_id}")
def update_question(
    question_id: int,
    question_data: schemas.AssignmentQuestionUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_admin)
):
    """Обновить вопрос"""
    question = db.query(models.AssignmentQuestion).filter(
        models.AssignmentQuestion.id == question_id
    ).first()
    if not question:
        raise HTTPException(status_code=404, detail="Question not found")
    
    update_data = question_data.model_dump(exclude_unset=True)
    if 'options' in update_data and update_data['options'] is not None:
        update_data['options'] = json.dumps(update_data['options'])
    
    for key, value in update_data.items():
        setattr(question, key, value)
    
    db.commit()
    return {"message": "Question updated"}


@router.delete("/questions/{question_id}")
def delete_question(
    question_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_admin)
):
    """Удалить вопрос"""
    question = db.query(models.AssignmentQuestion).filter(
        models.AssignmentQuestion.id == question_id
    ).first()
    if not question:
        raise HTTPException(status_code=404, detail="Question not found")
    
    db.delete(question)
    db.commit()
    return {"message": "Question deleted"}


# ========== ОТВЕТЫ НА ЗАДАНИЯ ==========

@router.post("/assignments/{assignment_id}/submit")
def submit_assignment(
    assignment_id: int,
    submit_data: schemas.AssignmentSubmit,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Отправить ответы на задание"""
    assignment = db.query(models.LessonAssignment).filter(
        models.LessonAssignment.id == assignment_id
    ).first()
    if not assignment:
        raise HTTPException(status_code=404, detail="Assignment not found")
    
    if assignment.deadline and assignment.deadline < datetime.utcnow():
        raise HTTPException(status_code=400, detail="Assignment deadline has passed")
    
    if submit_data.is_retake and not assignment.allow_retake:
        raise HTTPException(status_code=403, detail="Retake is not allowed for this assignment")
    
    remaining = get_remaining_attempts(assignment_id, current_user.id, db)
    if remaining <= 0 and assignment.max_attempts > 0:
        raise HTTPException(status_code=400, detail="No attempts remaining")
    
    latest_submission = db.query(models.AssignmentSubmission).filter(
        and_(
            models.AssignmentSubmission.assignment_id == assignment_id,
            models.AssignmentSubmission.user_id == current_user.id,
            models.AssignmentSubmission.is_latest == True
        )
    ).first()
    
    if submit_data.is_retake and latest_submission:
        db.query(models.UserAnswer).filter(
            models.UserAnswer.submission_id == latest_submission.id
        ).delete()
        db.query(models.AssignmentAttempt).filter(
            models.AssignmentAttempt.submission_id == latest_submission.id
        ).delete()
        db.delete(latest_submission)
        db.commit()
        latest_submission = None
    
    current_attempt = 1
    
    if latest_submission:
        current_attempt = latest_submission.current_attempt + 1
        latest_submission.is_latest = False
        db.commit()
    
    submission = models.AssignmentSubmission(
        assignment_id=assignment_id,
        user_id=current_user.id,
        current_attempt=current_attempt,
        is_latest=True,
        is_retake=submit_data.is_retake,
        retake_number=submit_data.is_retake and latest_submission and latest_submission.retake_number is not None and latest_submission.retake_number + 1 or 0,
        previous_submission_id=latest_submission.id if latest_submission else None,
        time_spent_seconds=submit_data.time_spent_seconds,
        can_retake=assignment.allow_retake
    )
    db.add(submission)
    db.flush()
    
    attempt = models.AssignmentAttempt(
        submission_id=submission.id,
        user_id=current_user.id,
        attempt_number=current_attempt,
        started_at=datetime.utcnow(),
        is_retake=submit_data.is_retake,
        time_spent_seconds=submit_data.time_spent_seconds
    )
    db.add(attempt)
    db.flush()
    
    total_points = 0
    has_text_answers = False
    has_file_answers = False
    auto_graded_count = 0
    
    for ans_data in submit_data.answers:
        question = db.query(models.AssignmentQuestion).filter(
            models.AssignmentQuestion.id == ans_data.question_id
        ).first()
        
        if not question:
            continue
        
        existing_answer = db.query(models.UserAnswer).filter(
            and_(
                models.UserAnswer.submission_id == submission.id,
                models.UserAnswer.question_id == question.id
            )
        ).first()
        
        if question.question_type == "text":
            has_text_answers = True
            if existing_answer:
                existing_answer.answer_text = ans_data.answer_text
                existing_answer.is_correct = None
                existing_answer.points_earned = 0
                existing_answer.attempt_id = attempt.id
            else:
                user_answer = models.UserAnswer(
                    submission_id=submission.id,
                    question_id=question.id,
                    answer_text=ans_data.answer_text,
                    is_correct=None,
                    points_earned=0,
                    attempt_id=attempt.id
                )
                db.add(user_answer)
                
        elif question.question_type == "choice":
            is_correct = False
            if question.correct_answer and ans_data.answer_text:
                try:
                    correct_idx = int(question.correct_answer)
                    user_idx = int(ans_data.answer_text)
                    is_correct = user_idx == correct_idx
                except:
                    pass
            
            points_earned = question.points if is_correct else 0
            total_points += points_earned
            auto_graded_count += 1
            
            if existing_answer:
                existing_answer.answer_text = ans_data.answer_text
                existing_answer.is_correct = is_correct
                existing_answer.points_earned = points_earned
                existing_answer.attempt_id = attempt.id
            else:
                user_answer = models.UserAnswer(
                    submission_id=submission.id,
                    question_id=question.id,
                    answer_text=ans_data.answer_text,
                    is_correct=is_correct,
                    points_earned=points_earned,
                    attempt_id=attempt.id
                )
                db.add(user_answer)
                
        elif question.question_type == "multiple":
            correct_answers = []
            if question.correct_answer:
                try:
                    correct_answers = json.loads(question.correct_answer)
                except:
                    pass
            
            user_answers = []
            if ans_data.answer_text:
                try:
                    user_answers = json.loads(ans_data.answer_text)
                except:
                    pass
            
            if isinstance(correct_answers, list) and isinstance(user_answers, list):
                correct_count = len(set(user_answers) & set(correct_answers))
                incorrect_count = len(set(user_answers) - set(correct_answers))
                
                if incorrect_count == 0 and correct_count == len(correct_answers):
                    is_correct = True
                    points_earned = question.points
                elif correct_count > 0:
                    is_correct = None
                    points_earned = int(question.points * (correct_count / len(correct_answers)))
                else:
                    is_correct = False
                    points_earned = 0
            else:
                is_correct = False
                points_earned = 0
            
            total_points += points_earned
            auto_graded_count += 1
            
            if existing_answer:
                existing_answer.answer_text = ans_data.answer_text
                existing_answer.is_correct = is_correct
                existing_answer.points_earned = points_earned
                existing_answer.attempt_id = attempt.id
            else:
                user_answer = models.UserAnswer(
                    submission_id=submission.id,
                    question_id=question.id,
                    answer_text=ans_data.answer_text,
                    is_correct=is_correct,
                    points_earned=points_earned,
                    attempt_id=attempt.id
                )
                db.add(user_answer)
                
        elif question.question_type == "file":
            has_file_answers = True
            if existing_answer:
                existing_answer.answer_file = ans_data.answer_file
                existing_answer.attempt_id = attempt.id
            else:
                user_answer = models.UserAnswer(
                    submission_id=submission.id,
                    question_id=question.id,
                    answer_file=ans_data.answer_file,
                    attempt_id=attempt.id
                )
                db.add(user_answer)
    
    submission.score = total_points
    submission.time_spent_seconds = submit_data.time_spent_seconds
    
    if not has_text_answers and not has_file_answers:
        submission.is_passed = total_points >= assignment.passing_score
    else:
        submission.is_passed = False
    
    attempt.completed_at = datetime.utcnow()
    attempt.score = total_points
    attempt.is_passed = submission.is_passed
    
    db.commit()
    
    if has_text_answers or has_file_answers:
        admins = db.query(models.User).filter(models.User.role == models.UserRole.ADMIN).all()
        for admin in admins:
            notification = models.Notification(
                user_id=admin.id,
                title="Новые ответы на проверку",
                message=f"Студент {current_user.full_name} отправил ответы на задание \"{assignment.title}\" в курсе \"{assignment.lesson.module.course.title}\". Требуется проверка."
            )
            db.add(notification)
        db.commit()
    
    if has_text_answers or has_file_answers:
        message = "Ответы отправлены на проверку учителю. После проверки вы сможете завершить урок."
    else:
        if submission.is_passed:
            message = f"Задание пройдено! Набрано {total_points} из {assignment.passing_score} баллов."
            notification = models.Notification(
                user_id=current_user.id,
                title="✅ Задание пройдено!",
                message=f"Вы успешно прошли задание \"{assignment.title}\" с результатом {total_points} из {assignment.passing_score} баллов."
            )
            db.add(notification)
            db.commit()
        else:
            remaining_attempts = get_remaining_attempts(assignment_id, current_user.id, db)
            if remaining_attempts > 0:
                message = f"Задание не пройдено. Набрано {total_points} из {assignment.passing_score} баллов. Осталось попыток: {remaining_attempts}"
            else:
                message = f"Задание не пройдено. Набрано {total_points} из {assignment.passing_score} баллов. Попыток больше нет."
    
    return {
        "message": message,
        "score": total_points,
        "max_score": assignment.max_score,
        "passed": submission.is_passed,
        "has_text_answers": has_text_answers,
        "has_file_answers": has_file_answers,
        "auto_graded_count": auto_graded_count,
        "attempt": submission.current_attempt,
        "max_attempts": assignment.max_attempts,
        "remaining_attempts": get_remaining_attempts(assignment_id, current_user.id, db),
        "is_retake": submit_data.is_retake,
        "can_retake": assignment.allow_retake
    }


# ========== ОТВЕТЫ ФАЙЛАМИ ==========

@router.post("/questions/{question_id}/file-answer")
async def answer_question_with_file(
    question_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Ответ на вопрос файлом"""
    question = db.query(models.AssignmentQuestion).filter(
        models.AssignmentQuestion.id == question_id
    ).first()
    if not question:
        raise HTTPException(status_code=404, detail="Question not found")
    
    if question.question_type != "file":
        raise HTTPException(status_code=400, detail="This question does not accept file answers")
    
    content = await file.read()
    file_size = len(content)
    
    if file_size > 50 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large (max 50MB)")
    
    submission = db.query(models.AssignmentSubmission).filter(
        and_(
            models.AssignmentSubmission.assignment_id == question.assignment_id,
            models.AssignmentSubmission.user_id == current_user.id,
            models.AssignmentSubmission.is_latest == True
        )
    ).first()
    
    if not submission:
        submission = models.AssignmentSubmission(
            assignment_id=question.assignment_id,
            user_id=current_user.id,
            current_attempt=1,
            is_latest=True
        )
        db.add(submission)
        db.flush()
    
    ext = file.filename.split('.')[-1].lower()
    filename = f"{uuid.uuid4().hex}.{ext}"
    filepath = os.path.join(ANSWERS_DIR, filename)
    
    with open(filepath, "wb") as buffer:
        buffer.write(content)
    
    file_url = f"/static/uploads/lms/answers/{filename}"
    file_type = get_file_type(file.filename)
    
    user_answer = db.query(models.UserAnswer).filter(
        and_(
            models.UserAnswer.submission_id == submission.id,
            models.UserAnswer.question_id == question_id
        )
    ).first()
    
    if user_answer:
        if user_answer.answer_file:
            old_file = user_answer.answer_file.replace("/static/uploads/lms/answers/", "")
            old_path = os.path.join(ANSWERS_DIR, old_file)
            if os.path.exists(old_path):
                os.remove(old_path)
        
        user_answer.answer_file = file_url
        user_answer.is_correct = None
        user_answer.points_earned = 0
    else:
        user_answer = models.UserAnswer(
            submission_id=submission.id,
            question_id=question_id,
            answer_file=file_url,
            is_correct=None,
            points_earned=0
        )
        db.add(user_answer)
    
    db.commit()
    
    return {
        "message": "File uploaded successfully",
        "file_url": file_url,
        "filename": file.filename,
        "file_size": file_size
    }


@router.get("/assignments/{assignment_id}/status")
def get_assignment_status(
    assignment_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Получить статус выполнения задания"""
    assignment = db.query(models.LessonAssignment).filter(
        models.LessonAssignment.id == assignment_id
    ).first()
    if not assignment:
        raise HTTPException(status_code=404, detail="Assignment not found")
    
    submission = db.query(models.AssignmentSubmission).filter(
        and_(
            models.AssignmentSubmission.assignment_id == assignment_id,
            models.AssignmentSubmission.user_id == current_user.id,
            models.AssignmentSubmission.is_latest == True
        )
    ).first()
    
    if not submission:
        return {
            "status": "not_started",
            "message": "Assignment not started",
            "score": None,
            "max_score": assignment.max_score,
            "passing_score": assignment.passing_score,
            "is_passed": False,
            "attempt": 0,
            "max_attempts": assignment.max_attempts,
            "submitted_at": None,
            "graded_at": None,
            "teacher_comment": None,
            "questions": [],
            "allow_retake": assignment.allow_retake,
            "retake_count": assignment.retake_count,
            "deadline": assignment.deadline,
            "is_overdue": False,
            "is_retake": False,
            "retake_number": 0,
            "can_retake": True,
            "remaining_attempts": assignment.max_attempts if assignment.max_attempts > 0 else 999
        }
    
    questions = db.query(models.AssignmentQuestion).filter(
        models.AssignmentQuestion.assignment_id == assignment_id
    ).all()
    
    question_status = []
    all_answered = True
    has_pending = False
    
    for question in questions:
        answer = db.query(models.UserAnswer).filter(
            and_(
                models.UserAnswer.submission_id == submission.id,
                models.UserAnswer.question_id == question.id
            )
        ).first()
        
        status_item = {
            "question_id": question.id,
            "question_type": question.question_type,
            "is_answered": answer is not None,
            "is_correct": answer.is_correct if answer else None,
            "points_earned": answer.points_earned if answer else 0,
            "answer_text": answer.answer_text if answer else None,
            "answer_file": answer.answer_file if answer else None,
            "hint": question.hint,
            "explanation": question.explanation
        }
        
        if answer and answer.is_correct is None and (answer.answer_text or answer.answer_file):
            status_item["status"] = "pending"
            has_pending = True
        elif answer and answer.is_correct is not None:
            status_item["status"] = "graded"
        elif not answer:
            status_item["status"] = "not_answered"
            all_answered = False
        else:
            status_item["status"] = "unknown"
        
        question_status.append(status_item)
    
    is_overdue = False
    if assignment.deadline and assignment.deadline < datetime.utcnow():
        is_overdue = True
    
    can_retake = assignment.allow_retake
    remaining_attempts = get_remaining_attempts(assignment_id, current_user.id, db)
    if assignment.max_attempts > 0 and remaining_attempts <= 0:
        can_retake = False
    
    if submission.is_passed:
        overall_status = "passed"
        message = f"✅ Задание пройдено! Баллы: {submission.score}/{assignment.max_score}"
    elif has_pending:
        overall_status = "pending_teacher"
        message = "⏳ Ожидает проверки учителем"
    elif submission.score is not None and submission.score < assignment.passing_score:
        if submission.teacher_comment and "пересдачу" in submission.teacher_comment.lower():
            overall_status = "retake"
            message = f"🔄 Отправлено на пересдачу. Комментарий: {submission.teacher_comment}"
        elif can_retake and remaining_attempts > 0:
            overall_status = "failed"
            message = f"❌ Не пройдено. Баллы: {submission.score}/{assignment.passing_score}. Осталось попыток: {remaining_attempts}"
        else:
            overall_status = "failed"
            message = f"❌ Не пройдено. Баллы: {submission.score}/{assignment.passing_score}. Попыток больше нет."
    elif not all_answered:
        overall_status = "in_progress"
        message = "✏️ Ответьте на все вопросы"
    else:
        overall_status = "submitted"
        message = "📤 Отправлено на проверку"
    
    return {
        "status": overall_status,
        "message": message,
        "score": submission.score,
        "max_score": assignment.max_score,
        "passing_score": assignment.passing_score,
        "is_passed": submission.is_passed,
        "attempt": submission.current_attempt,
        "max_attempts": assignment.max_attempts,
        "submitted_at": submission.submitted_at,
        "graded_at": submission.graded_at,
        "teacher_comment": submission.teacher_comment,
        "questions": question_status,
        "allow_retake": assignment.allow_retake,
        "retake_count": assignment.retake_count,
        "deadline": assignment.deadline,
        "is_overdue": is_overdue,
        "is_retake": submission.is_retake,
        "retake_number": submission.retake_number,
        "can_retake": can_retake,
        "remaining_attempts": remaining_attempts
    }


# ========== ИНФОРМАЦИЯ О ПОПЫТКАХ ==========

@router.get("/assignments/{assignment_id}/attempts-info")
def get_attempts_info(
    assignment_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """Получить информацию о попытках выполнения задания"""
    assignment = db.query(models.LessonAssignment).filter(
        models.LessonAssignment.id == assignment_id
    ).first()
    if not assignment:
        raise HTTPException(status_code=404, detail="Assignment not found")
    
    submissions = db.query(models.AssignmentSubmission).filter(
        and_(
            models.AssignmentSubmission.assignment_id == assignment_id,
            models.AssignmentSubmission.user_id == current_user.id
        )
    ).order_by(models.AssignmentSubmission.current_attempt.desc()).all()
    
    return {
        "assignment_id": assignment_id,
        "max_attempts": assignment.max_attempts,
        "allow_retake": assignment.allow_retake,
        "retake_count": assignment.retake_count,
        "total_attempts": len(submissions),
        "remaining_attempts": get_remaining_attempts(assignment_id, current_user.id, db),
        "attempts": [{
            "attempt_number": s.current_attempt,
            "score": s.score,
            "is_passed": s.is_passed,
            "is_retake": s.is_retake,
            "submitted_at": s.submitted_at,
            "graded_at": s.graded_at
        } for s in submissions]
    }


# ========== УЧИТЕЛЬСКИЙ ЖУРНАЛ ==========

@router.get("/courses/{course_id}/teacher/journal")
def get_teacher_journal(
    course_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_admin)
):
    """Получить журнал успеваемости для учителя"""
    course = db.query(models.Course).filter(models.Course.id == course_id).first()
    if not course:
        raise HTTPException(status_code=404, detail="Course not found")
    
    registrations = db.query(models.CourseRegistration, models.User).join(
        models.User
    ).filter(models.CourseRegistration.course_id == course_id).all()
    
    modules = db.query(models.CourseModule).filter(
        models.CourseModule.course_id == course_id
    ).order_by(models.CourseModule.order_index).all()
    
    result = {
        "course": {"id": course.id, "title": course.title, "format_type": course.format_type},
        "modules": [{"id": m.id, "title": m.title} for m in modules],
        "students": []
    }
    
    for reg, user in registrations:
        student_data = {
            "user_id": user.id,
            "full_name": user.full_name,
            "email": user.email,
            "position": user.position,
            "organization": user.organization,
            "modules": [],
            "total_progress": 0,
            "is_completed": False,
            "retake_count": 0
        }
        
        total_lessons = 0
        completed_lessons = 0
        retakes = 0
        
        for module in modules:
            lessons = db.query(models.CourseLesson).filter(
                models.CourseLesson.module_id == module.id
            ).all()
            
            module_total = len(lessons)
            module_completed = 0
            
            for lesson in lessons:
                total_lessons += 1
                progress = db.query(models.UserLessonProgress).filter(
                    and_(
                        models.UserLessonProgress.user_id == user.id,
                        models.UserLessonProgress.lesson_id == lesson.id,
                        models.UserLessonProgress.is_completed == True
                    )
                ).first()
                if progress:
                    completed_lessons += 1
                    module_completed += 1
            
            module_progress = db.query(models.UserModuleProgress).filter(
                and_(
                    models.UserModuleProgress.user_id == user.id,
                    models.UserModuleProgress.module_id == module.id
                )
            ).first()
            
            if lesson and lesson.assignment:
                retakes_in_module = db.query(models.AssignmentSubmission).filter(
                    and_(
                        models.AssignmentSubmission.assignment_id == lesson.assignment.id,
                        models.AssignmentSubmission.user_id == user.id,
                        models.AssignmentSubmission.is_retake == True
                    )
                ).count()
                retakes += retakes_in_module
            
            student_data["modules"].append({
                "module_id": module.id,
                "is_completed": module_progress.is_completed if module_progress else False,
                "completed_by_teacher": module_progress.completed_by_teacher if module_progress else False,
                "completed_at": module_progress.completed_at if module_progress else None,
                "teacher_comment": module_progress.teacher_comment if module_progress else None,
                "progress_percent": int((module_completed / module_total) * 100) if module_total > 0 else 0
            })
        
        student_data["total_progress"] = int((completed_lessons / total_lessons) * 100) if total_lessons > 0 else 0
        student_data["is_completed"] = student_data["total_progress"] == 100
        student_data["retake_count"] = retakes
        
        result["students"].append(student_data)
    
    return result


@router.post("/teacher/mark-module-completed")
def mark_module_completed(
    data: dict,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_admin)
):
    """Отметить модуль как пройденный для студента"""
    user_id = data.get("user_id")
    module_id = data.get("module_id")
    comment = data.get("comment", "")
    
    module = db.query(models.CourseModule).filter(models.CourseModule.id == module_id).first()
    if not module:
        raise HTTPException(status_code=404, detail="Module not found")
    
    module_progress = db.query(models.UserModuleProgress).filter(
        and_(
            models.UserModuleProgress.user_id == user_id,
            models.UserModuleProgress.module_id == module_id
        )
    ).first()
    
    if not module_progress:
        module_progress = models.UserModuleProgress(
            user_id=user_id,
            module_id=module_id,
            is_completed=True,
            completed_by_teacher=True,
            completed_at=datetime.utcnow(),
            teacher_comment=comment
        )
        db.add(module_progress)
    else:
        module_progress.is_completed = True
        module_progress.completed_by_teacher = True
        module_progress.completed_at = datetime.utcnow()
        module_progress.teacher_comment = comment
    
    db.commit()
    
    lessons = db.query(models.CourseLesson).filter(
        models.CourseLesson.module_id == module_id
    ).all()
    
    for lesson in lessons:
        lesson_progress = db.query(models.UserLessonProgress).filter(
            and_(
                models.UserLessonProgress.user_id == user_id,
                models.UserLessonProgress.lesson_id == lesson.id
            )
        ).first()
        
        if not lesson_progress:
            lesson_progress = models.UserLessonProgress(
                user_id=user_id,
                lesson_id=lesson.id,
                is_completed=True,
                completed_at=datetime.utcnow()
            )
            db.add(lesson_progress)
        elif not lesson_progress.is_completed:
            lesson_progress.is_completed = True
            lesson_progress.completed_at = datetime.utcnow()
    
    db.commit()
    
    update_course_progress(user_id, module.course_id, db)
    
    course_modules = db.query(models.CourseModule).filter(
        models.CourseModule.course_id == module.course_id
    ).all()
    
    completed_modules = db.query(models.UserModuleProgress).filter(
        and_(
            models.UserModuleProgress.user_id == user_id,
            models.UserModuleProgress.module_id.in_([m.id for m in course_modules]),
            models.UserModuleProgress.is_completed == True
        )
    ).count()
    
    if completed_modules == len(course_modules):
        course_progress = db.query(models.UserProgress).filter(
            and_(
                models.UserProgress.user_id == user_id,
                models.UserProgress.course_id == module.course_id
            )
        ).first()
        
        if course_progress:
            course_progress.is_completed = True
            course_progress.completed_at = datetime.utcnow()
            course_progress.progress_percent = 100
            db.commit()
            
            from app.routers.achievements_router import check_and_award_achievements
            check_and_award_achievements(user_id, db)
    
    return {"message": "Module marked as completed"}


# ========== ПРОВЕРКА ОТВЕТОВ ==========

@router.get("/courses/{course_id}/submissions/pending")
def get_pending_submissions(
    course_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_admin)
):
    """Получить все непроверенные текстовые ответы для курса"""
    course = db.query(models.Course).filter(models.Course.id == course_id).first()
    if not course:
        raise HTTPException(status_code=404, detail="Course not found")
    
    modules = db.query(models.CourseModule).filter(
        models.CourseModule.course_id == course_id
    ).all()
    
    module_ids = [m.id for m in modules]
    
    lessons = db.query(models.CourseLesson).filter(
        models.CourseLesson.module_id.in_(module_ids)
    ).all()
    
    lesson_ids = [l.id for l in lessons]
    
    assignments = db.query(models.LessonAssignment).filter(
        models.LessonAssignment.lesson_id.in_(lesson_ids)
    ).all()
    
    assignment_ids = [a.id for a in assignments]
    
    pending_questions = db.query(models.AssignmentQuestion).filter(
        and_(
            models.AssignmentQuestion.assignment_id.in_(assignment_ids),
            models.AssignmentQuestion.question_type.in_(['text', 'file'])
        )
    ).all()
    
    pending_submissions = []
    
    for question in pending_questions:
        if question.question_type == 'text':
            answers = db.query(models.UserAnswer).filter(
                and_(
                    models.UserAnswer.question_id == question.id,
                    models.UserAnswer.is_correct.is_(None),
                    models.UserAnswer.answer_text.isnot(None)
                )
            ).all()
        else:
            answers = db.query(models.UserAnswer).filter(
                and_(
                    models.UserAnswer.question_id == question.id,
                    models.UserAnswer.is_correct.is_(None),
                    models.UserAnswer.answer_file.isnot(None)
                )
            ).all()
        
        for answer in answers:
            submission = db.query(models.AssignmentSubmission).filter(
                models.AssignmentSubmission.id == answer.submission_id
            ).first()
            
            if submission:
                user = db.query(models.User).filter(models.User.id == submission.user_id).first()
                lesson = db.query(models.CourseLesson).filter(
                    models.CourseLesson.id == question.assignment.lesson_id
                ).first()
                
                if lesson and lesson.module.course_id == course_id:
                    pending_submissions.append({
                        "submission_id": submission.id,
                        "question_id": question.id,
                        "answer_text": answer.answer_text,
                        "answer_file": answer.answer_file,
                        "question_type": question.question_type,
                        "student_name": user.full_name if user else "Unknown",
                        "student_email": user.email if user else "Unknown",
                        "question_text": question.question_text,
                        "course_title": course.title,
                        "lesson_title": lesson.title,
                        "points": question.points,
                        "is_retake": submission.is_retake,
                        "retake_number": submission.retake_number
                    })
    
    return pending_submissions


@router.get("/submissions/{submission_id}")
def get_submission_details(
    submission_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_admin)
):
    """Получить детали submission для проверки"""
    submission = db.query(models.AssignmentSubmission).filter(
        models.AssignmentSubmission.id == submission_id
    ).first()
    
    if not submission:
        raise HTTPException(status_code=404, detail="Submission not found")
    
    user = db.query(models.User).filter(models.User.id == submission.user_id).first()
    assignment = submission.assignment
    lesson = assignment.lesson
    module = lesson.module
    course = module.course
    
    answers = []
    for answer in submission.answers:
        question = db.query(models.AssignmentQuestion).filter(
            models.AssignmentQuestion.id == answer.question_id
        ).first()
        
        if question:
            answers.append({
                "question_id": question.id,
                "question_text": question.question_text,
                "question_type": question.question_type,
                "answer_text": answer.answer_text,
                "answer_file": answer.answer_file,
                "is_correct": answer.is_correct,
                "points_earned": answer.points_earned,
                "max_points": question.points,
                "teacher_comment": answer.teacher_comment,
                "auto_graded": answer.auto_graded
            })
    
    return {
        "submission_id": submission.id,
        "student_name": user.full_name if user else "Unknown",
        "student_email": user.email if user else "Unknown",
        "course_title": course.title,
        "module_title": module.title,
        "lesson_title": lesson.title,
        "assignment_title": assignment.title,
        "submitted_at": submission.submitted_at,
        "score": submission.score,
        "max_score": assignment.max_score,
        "is_passed": submission.is_passed,
        "graded_by": submission.graded_by,
        "graded_at": submission.graded_at,
        "teacher_comment": submission.teacher_comment,
        "answers": answers,
        "is_retake": submission.is_retake,
        "retake_number": submission.retake_number,
        "previous_submission_id": submission.previous_submission_id,
        "time_spent_seconds": submission.time_spent_seconds,
        "attempts": [{
            "attempt_number": a.attempt_number,
            "started_at": a.started_at,
            "completed_at": a.completed_at,
            "score": a.score,
            "is_passed": a.is_passed,
            "is_retake": a.is_retake
        } for a in submission.attempts]
    }


@router.post("/grade-text-answer")
def grade_text_answer(
    data: schemas.GradeTextAnswer,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_admin)
):
    """Оценить текстовый ответ студента"""
    submission = db.query(models.AssignmentSubmission).filter(
        models.AssignmentSubmission.id == data.submission_id
    ).first()
    
    if not submission:
        raise HTTPException(status_code=404, detail="Submission not found")
    
    user_answer = db.query(models.UserAnswer).filter(
        and_(
            models.UserAnswer.submission_id == data.submission_id,
            models.UserAnswer.question_id == data.question_id
        )
    ).first()
    
    if not user_answer:
        raise HTTPException(status_code=404, detail="Answer not found")
    
    question = db.query(models.AssignmentQuestion).filter(
        models.AssignmentQuestion.id == data.question_id
    ).first()
    
    if not question:
        raise HTTPException(status_code=404, detail="Question not found")
    
    if data.status == "retake":
        user_answer.teacher_comment = data.comment
        db.commit()
        
        submission.teacher_comment = f"🔄 Отправлено на пересдачу. {data.comment}"
        submission.is_passed = False
        submission.is_retake = True
        db.commit()
        
        notification = models.Notification(
            user_id=submission.user_id,
            title="🔄 Ответ отправлен на пересдачу",
            message=f"Ваш ответ на вопрос \"{question.question_text[:50]}...\" отправлен на пересдачу. {data.comment}"
        )
        db.add(notification)
        db.commit()
        
        return {"message": "Answer marked for retake. Student can try again."}
    
    if data.status == "pass":
        user_answer.is_correct = True
        user_answer.points_earned = data.points
        user_answer.teacher_comment = data.comment
    elif data.status == "fail":
        user_answer.is_correct = False
        user_answer.points_earned = 0
        user_answer.teacher_comment = data.comment
    else:
        raise HTTPException(status_code=400, detail="Invalid status")
    
    db.commit()
    
    pending_answers = db.query(models.UserAnswer).filter(
        and_(
            models.UserAnswer.submission_id == data.submission_id,
            models.UserAnswer.is_correct.is_(None)
        )
    ).count()
    
    if pending_answers == 0:
        submission.graded_by = current_user.id
        submission.graded_at = datetime.utcnow()
        submission.teacher_comment = data.comment
        
        all_answers = db.query(models.UserAnswer).filter(
            models.UserAnswer.submission_id == data.submission_id
        ).all()
        
        total_points = sum(a.points_earned for a in all_answers if a.is_correct is not None)
        submission.score = total_points
        
        if total_points >= submission.assignment.passing_score:
            submission.is_passed = True
        else:
            submission.is_passed = False
        
        db.commit()
        
        if submission.is_passed:
            lesson = submission.assignment.lesson
            
            user_progress = db.query(models.UserLessonProgress).filter(
                and_(
                    models.UserLessonProgress.user_id == submission.user_id,
                    models.UserLessonProgress.lesson_id == lesson.id
                )
            ).first()
            
            if not user_progress:
                user_progress = models.UserLessonProgress(
                    user_id=submission.user_id,
                    lesson_id=lesson.id,
                    is_completed=True,
                    completed_at=datetime.utcnow()
                )
                db.add(user_progress)
            elif not user_progress.is_completed:
                user_progress.is_completed = True
                user_progress.completed_at = datetime.utcnow()
            
            db.commit()
            update_course_progress(submission.user_id, lesson.module.course_id, db)
    
    notification = models.Notification(
        user_id=submission.user_id,
        title="📝 Ответ проверен",
        message=f"Ваш ответ на вопрос \"{question.question_text[:50]}...\" проверен. " + (
            f"✅ Зачтено! Получено баллов: {data.points}" if data.status == "pass" else
            f"❌ Не зачтено. {data.comment}"
        )
    )
    db.add(notification)
    db.commit()
    
    return {
        "message": "Answer graded successfully",
        "status": data.status,
        "points": data.points if data.status == "pass" else 0,
        "comment": data.comment,
        "pending_answers_left": pending_answers
    }


# ========== ЭКСПОРТ РЕЗУЛЬТАТОВ ==========

@router.get("/courses/{course_id}/export-results")
def export_course_results(
    course_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_admin)
):
    """Экспорт результатов курса в Excel"""
    course = db.query(models.Course).filter(models.Course.id == course_id).first()
    if not course:
        raise HTTPException(status_code=404, detail="Course not found")
    
    registrations = db.query(models.CourseRegistration, models.User).join(
        models.User
    ).filter(models.CourseRegistration.course_id == course_id).all()
    
    modules = db.query(models.CourseModule).filter(
        models.CourseModule.course_id == course_id
    ).all()
    
    all_lessons = []
    for module in modules:
        lessons = db.query(models.CourseLesson).filter(
            models.CourseLesson.module_id == module.id
        ).all()
        all_lessons.extend(lessons)
    
    data = []
    for reg, user in registrations:
        submissions = db.query(models.AssignmentSubmission).filter(
            models.AssignmentSubmission.user_id == user.id
        ).all()
        
        total_score = sum(s.score or 0 for s in submissions)
        passed_count = sum(1 for s in submissions if s.is_passed)
        total_assignments = len(submissions)
        retake_count = sum(1 for s in submissions if s.is_retake)
        
        completed_lessons = db.query(models.UserLessonProgress).filter(
            and_(
                models.UserLessonProgress.user_id == user.id,
                models.UserLessonProgress.is_completed == True
            )
        ).count()
        
        total_lessons = len(all_lessons)
        progress_percent = int((completed_lessons / total_lessons) * 100) if total_lessons > 0 else 0
        
        data.append({
            "full_name": user.full_name,
            "email": user.email,
            "position": user.position or "",
            "organization": user.organization or "",
            "total_score": total_score,
            "passed_assignments": passed_count,
            "total_assignments": total_assignments,
            "retake_count": retake_count,
            "completed_lessons": completed_lessons,
            "total_lessons": total_lessons,
            "progress_percent": progress_percent,
            "status": "Завершён" if progress_percent == 100 else "В процессе",
            "started_at": user.created_at.strftime("%d.%m.%Y") if user.created_at else "",
            "completed_at": ""
        })
        
        if progress_percent == 100:
            course_progress = db.query(models.UserProgress).filter(
                and_(
                    models.UserProgress.user_id == user.id,
                    models.UserProgress.course_id == course_id
                )
            ).first()
            if course_progress and course_progress.completed_at:
                data[-1]["completed_at"] = course_progress.completed_at.strftime("%d.%m.%Y")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Результаты"
    
    headers = ["№", "ФИО", "Email", "Должность", "Организация", "Баллов", "Пройдено заданий", "Всего заданий", 
               "Пересдач", "Пройдено уроков", "Всего уроков", "Прогресс", "Статус", "Дата начала", "Дата завершения"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0057A4", end_color="0057A4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for row, item in enumerate(data, 2):
        ws.cell(row=row, column=1, value=row - 1)
        ws.cell(row=row, column=2, value=item["full_name"])
        ws.cell(row=row, column=3, value=item["email"])
        ws.cell(row=row, column=4, value=item["position"])
        ws.cell(row=row, column=5, value=item["organization"])
        ws.cell(row=row, column=6, value=item["total_score"])
        ws.cell(row=row, column=7, value=item["passed_assignments"])
        ws.cell(row=row, column=8, value=item["total_assignments"])
        ws.cell(row=row, column=9, value=item["retake_count"])
        ws.cell(row=row, column=10, value=item["completed_lessons"])
        ws.cell(row=row, column=11, value=item["total_lessons"])
        ws.cell(row=row, column=12, value=f"{item['progress_percent']}%")
        ws.cell(row=row, column=13, value=item["status"])
        ws.cell(row=row, column=14, value=item["started_at"])
        ws.cell(row=row, column=15, value=item["completed_at"])
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 20
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"course_results_{course_id}_{timestamp}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename}",
            "Access-Control-Expose-Headers": "Content-Disposition"
        }
    )


# ========== СТАТИСТИКА ==========

@router.get("/courses/{course_id}/statistics")
def get_course_statistics(
    course_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_admin)
):
    """Получить статистику по курсу"""
    course = db.query(models.Course).filter(models.Course.id == course_id).first()
    if not course:
        raise HTTPException(status_code=404, detail="Course not found")
    
    total_students = db.query(models.CourseRegistration).filter(
        models.CourseRegistration.course_id == course_id
    ).count()
    
    active_students = db.query(models.UserProgress).filter(
        and_(
            models.UserProgress.course_id == course_id,
            models.UserProgress.progress_percent > 0,
            models.UserProgress.progress_percent < 100
        )
    ).count()
    
    completed_students = db.query(models.UserProgress).filter(
        and_(
            models.UserProgress.course_id == course_id,
            models.UserProgress.is_completed == True
        )
    ).count()
    
    avg_progress = db.query(func.avg(models.UserProgress.progress_percent)).filter(
        models.UserProgress.course_id == course_id
    ).scalar() or 0
    
    modules = db.query(models.CourseModule).filter(
        models.CourseModule.course_id == course_id
    ).all()
    module_ids = [m.id for m in modules]
    
    lessons = db.query(models.CourseLesson).filter(
        models.CourseLesson.module_id.in_(module_ids)
    ).all()
    lesson_ids = [l.id for l in lessons]
    
    assignments = db.query(models.LessonAssignment).filter(
        models.LessonAssignment.lesson_id.in_(lesson_ids)
    ).all()
    
    assignment_stats = []
    for assignment in assignments:
        submissions = db.query(models.AssignmentSubmission).filter(
            models.AssignmentSubmission.assignment_id == assignment.id
        ).all()
        
        total_subs = len(submissions)
        avg_score = sum(s.score or 0 for s in submissions) / total_subs if total_subs > 0 else 0
        passed = sum(1 for s in submissions if s.is_passed)
        retakes = sum(1 for s in submissions if s.is_retake)
        
        assignment_stats.append({
            "assignment_id": assignment.id,
            "title": assignment.title,
            "total_submissions": total_subs,
            "avg_score": round(avg_score, 2),
            "passed_count": passed,
            "pass_rate": round((passed / total_subs * 100), 2) if total_subs > 0 else 0,
            "retake_count": retakes
        })
    
    return {
        "course_id": course_id,
        "course_title": course.title,
        "total_students": total_students,
        "active_students": active_students,
        "completed_students": completed_students,
        "average_progress": round(avg_progress, 2),
        "total_assignments": len(assignments),
        "assignments": assignment_stats,
        "completion_rate": round((completed_students / total_students * 100), 2) if total_students > 0 else 0
    }