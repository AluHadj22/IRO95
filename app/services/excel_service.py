# app/services/excel_service.py
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO
from typing import List, Dict
from sqlalchemy.orm import Session, joinedload, selectinload
from app import models
from app.services.encryption_service import EncryptionService
import json
import re
import logging
from datetime import datetime

logger = logging.getLogger(__name__)

encryption = EncryptionService()


def escape_excel_string(value) -> str:
    if value is None:
        return ""
    
    if not isinstance(value, str):
        return str(value)
    
    if not value:
        return ""
    
    if value and value[0] in '=+-@':
        return "'" + value
    
    value = value.replace('<', '&lt;').replace('>', '&gt;')
    
    return value


def sanitize_snils(snils: str) -> str:
    if not snils:
        return ""
    
    try:
        decrypted = encryption.decrypt(snils)
        if decrypted:
            return decrypted
    except Exception:
        pass
    
    return snils


def sanitize_phone(phone: str) -> str:
    if not phone:
        return ""
    
    cleaned = re.sub(r'[^0-9+]', '', phone)
    
    if cleaned.startswith('8') and len(cleaned) == 11:
        cleaned = '+7' + cleaned[1:]
    
    if not cleaned.startswith('+') and len(cleaned) >= 10:
        if cleaned.startswith('7'):
            cleaned = '+' + cleaned
        else:
            cleaned = '+7' + cleaned
    
    return cleaned


def get_column_letter(col_idx: int) -> str:
    result = ""
    while col_idx > 0:
        col_idx -= 1
        result = chr(65 + col_idx % 26) + result
        col_idx //= 26
    return result


def generate_registrations_excel(data: List[Dict], course_title: str = "") -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Регистрации"
    
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0057A4', end_color='0057A4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    cell_alignment = Alignment(horizontal='left', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    if course_title:
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = escape_excel_string(course_title)
        title_cell.font = Font(name='Arial', size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center')
    
    headers = ["№", "ФИО", "Email", "Телефон", "Должность", "Организация"]
    start_row = 3 if course_title else 1
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=escape_excel_string(header))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for idx, row in enumerate(data, start=start_row + 1):
        ws.cell(row=idx, column=1, value=idx - start_row).border = thin_border
        ws.cell(row=idx, column=2, value=escape_excel_string(row.get('full_name', ''))).border = thin_border
        ws.cell(row=idx, column=3, value=escape_excel_string(row.get('email', ''))).border = thin_border
        ws.cell(row=idx, column=4, value=escape_excel_string(row.get('phone', ''))).border = thin_border
        ws.cell(row=idx, column=5, value=escape_excel_string(row.get('position', ''))).border = thin_border
        ws.cell(row=idx, column=6, value=escape_excel_string(row.get('organization', ''))).border = thin_border
    
    column_widths = [5, 30, 30, 15, 25, 35]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    ws.freeze_panes = f'A{start_row + 1}'
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _get_users_export_data_batch(db: Session, users: List[models.User]) -> Dict[int, Dict[str, str]]:
    if not users:
        return {}
    
    user_ids = [u.id for u in users]
    
    users_with_data = db.query(models.User).options(
        selectinload(models.User.work),
        selectinload(models.User.education),
        selectinload(models.User.additional_info)
    ).filter(models.User.id.in_(user_ids)).all()
    
    users_map = {u.id: u for u in users_with_data}
    
    result = {}
    
    for user in users:
        user_data = users_map.get(user.id)
        if not user_data:
            continue
        
        current_work = None
        if user_data.work:
            for w in user_data.work:
                if w.is_current:
                    current_work = w
                    break
            if not current_work:
                current_work = user_data.work[0] if user_data.work else None
        
        main_education = None
        if user_data.education:
            for e in user_data.education:
                if e.is_main:
                    main_education = e
                    break
            if not main_education:
                main_education = user_data.education[0] if user_data.education else None
        
        additional_info = user_data.additional_info
        
        subjects = []
        if current_work and current_work.subjects:
            try:
                subjects = json.loads(current_work.subjects)
                if not isinstance(subjects, list):
                    subjects = []
            except (json.JSONDecodeError, TypeError):
                subjects = []
        
        birth_date_str = ""
        if user.birth_date:
            try:
                birth_date_str = user.birth_date.strftime("%d.%m.%Y")
            except Exception:
                birth_date_str = ""
        
        gender_str = ""
        if user.gender == "male":
            gender_str = "Мужской"
        elif user.gender == "female":
            gender_str = "Женский"
        
        snils_value = ""
        if additional_info and additional_info.snils:
            snils_value = sanitize_snils(additional_info.snils)
        
        phone_value = user.phone_raw or user.phone or ""
        phone_value = sanitize_phone(phone_value)
        
        result[user.id] = {
            "municipality": escape_excel_string(user.municipality),
            "organization": escape_excel_string(current_work.organization if current_work else ""),
            "last_name": escape_excel_string(user.last_name),
            "first_name": escape_excel_string(user.first_name),
            "middle_name": escape_excel_string(user.middle_name),
            "birth_date": escape_excel_string(birth_date_str),
            "gender": escape_excel_string(gender_str),
            "snils": escape_excel_string(snils_value),
            "position": escape_excel_string(current_work.position if current_work else ""),
            "activity_type": escape_excel_string(current_work.activity_type if current_work else ""),
            "subjects": escape_excel_string("; ".join(subjects) if subjects else ""),
            "education": escape_excel_string(main_education.education_level if main_education else ""),
            "document_series": escape_excel_string(main_education.document_series if main_education else ""),
            "document_number": escape_excel_string(main_education.document_number if main_education else ""),
            "teaching_experience": current_work.teaching_experience_years if current_work and current_work.teaching_experience_years is not None else "",
            "work_experience": current_work.work_experience_years if current_work and current_work.work_experience_years is not None else "",
            "email": escape_excel_string(user.email),
            "phone": escape_excel_string(phone_value)
        }
    
    return result


def generate_full_registrations_excel(
    db: Session, 
    course_id: int, 
    course_title: str = ""
) -> BytesIO:
    try:
        registrations = db.query(models.CourseRegistration, models.User).join(
            models.User
        ).filter(models.CourseRegistration.course_id == course_id).order_by(
            models.CourseRegistration.registered_at.desc()
        ).all()
        
        if not registrations:
            return generate_empty_registrations_excel(course_title)
        
        users = [reg.User for reg in registrations]
        
        users_data = _get_users_export_data_batch(db, users)
        
        data = []
        for idx, (reg, user) in enumerate(registrations, 1):
            user_data = users_data.get(user.id, {})
            data.append({
                "number": idx,
                "municipality": user_data.get("municipality", ""),
                "organization": user_data.get("organization", ""),
                "last_name": user_data.get("last_name", ""),
                "first_name": user_data.get("first_name", ""),
                "middle_name": user_data.get("middle_name", ""),
                "birth_date": user_data.get("birth_date", ""),
                "gender": user_data.get("gender", ""),
                "snils": user_data.get("snils", ""),
                "position": user_data.get("position", ""),
                "activity_type": user_data.get("activity_type", ""),
                "subjects": user_data.get("subjects", ""),
                "education": user_data.get("education", ""),
                "document_series": user_data.get("document_series", ""),
                "document_number": user_data.get("document_number", ""),
                "teaching_experience": user_data.get("teaching_experience", ""),
                "work_experience": user_data.get("work_experience", ""),
                "email": user_data.get("email", ""),
                "phone": user_data.get("phone", "")
            })
        
        logger.info(f"Exported {len(data)} registrations for course {course_id}")
        return generate_full_registrations_excel_from_data(data, course_title)
        
    except Exception as e:
        logger.error(f"Error exporting registrations: {str(e)}")
        raise


def generate_full_registrations_excel_from_data(data: List[Dict], course_title: str = "") -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Регистрации"
    
    header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0057A4', end_color='0057A4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    start_row = 1
    
    if course_title:
        ws.merge_cells('A1:S1')
        title_cell = ws['A1']
        title_cell.value = escape_excel_string(f"Регистрации на курс: {course_title}")
        title_cell.font = Font(name='Arial', size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        start_row = 3
    
    headers = [
        "№", "Муниципалитет", "Место работы", "Фамилия", "Имя", "Отчество",
        "Дата рождения", "Пол", "СНИЛС", "Должность", "Вид деятельности",
        "Предмет", "Образование", "Серия диплома", "Номер диплома",
        "Педагогический стаж", "Стаж в должности", "Email", "Телефон"
    ]
    
    col_indices = {
        "number": 1,
        "municipality": 2,
        "organization": 3,
        "last_name": 4,
        "first_name": 5,
        "middle_name": 6,
        "birth_date": 7,
        "gender": 8,
        "snils": 9,
        "position": 10,
        "activity_type": 11,
        "subjects": 12,
        "education": 13,
        "document_series": 14,
        "document_number": 15,
        "teaching_experience": 16,
        "work_experience": 17,
        "email": 18,
        "phone": 19
    }
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=escape_excel_string(header))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row_idx, row_data in enumerate(data, start=start_row + 1):
        for field, col_idx in col_indices.items():
            value = row_data.get(field, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border
    
    column_widths = {
        1: 5, 2: 30, 3: 40, 4: 20, 5: 20, 6: 20,
        7: 18, 8: 15, 9: 20, 10: 25, 11: 25, 12: 30,
        13: 25, 14: 18, 15: 18, 16: 18, 17: 18, 18: 30, 19: 18
    }
    
    for col_idx, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    ws.row_dimensions[start_row].height = 35
    for row_idx in range(start_row + 1, start_row + 1 + len(data)):
        ws.row_dimensions[row_idx].height = 25
    
    ws.freeze_panes = f'A{start_row + 1}'
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_empty_registrations_excel(course_title: str = "") -> BytesIO:
    return generate_full_registrations_excel_from_data([], course_title)