# app/services/excel_export_service.py
import os
import json
import re
import logging
from io import BytesIO
from datetime import datetime
from typing import List, Dict, Any, Optional
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, joinedload
from app import models
from app.services.encryption_service import EncryptionService

logger = logging.getLogger(__name__)

encryption = EncryptionService()


class ExcelExportService:
    """Сервис для экспорта данных пользователей в Excel по шаблону"""
    
    TEMPLATE_PATH = "app/static/templates/export_template.xlsx"
    
    COLUMN_MAPPING = {
        "municipality": 3,
        "organization": 4,
        "last_name": 5,
        "first_name": 6,
        "middle_name": 7,
        "birth_date": 8,
        "gender": 9,
        "snils": 10,
        "position": 11,
        "activity_type": 12,
        "subjects": 13,
        "education": 14,
        "document_series": 15,
        "document_number": 16,
        "teaching_experience": 17,
        "work_experience": 18,
        "email": 19,
        "phone": 20
    }
    
    def __init__(self, db: Session):
        self.db = db
    
    def _escape_excel_string(self, value: Any) -> str:
        if value is None:
            return ""
        
        if not isinstance(value, str):
            return str(value)
        
        if not value:
            return ""
        
        if value and value[0] in '=+-@':
            return "'" + value
        
        value = value.replace('<', '&lt;').replace('>', '&gt;')
        
        return value
    
    def _sanitize_snils(self, snils: str) -> str:
        if not snils:
            return ""
        
        try:
            decrypted = encryption.decrypt(snils)
            if decrypted:
                return decrypted
        except Exception:
            pass
        
        return snils
    
    def _sanitize_phone(self, phone: str) -> str:
        if not phone:
            return ""
        
        cleaned = re.sub(r'[^0-9+]', '', phone)
        
        if cleaned.startswith('8') and len(cleaned) == 11:
            cleaned = '+7' + cleaned[1:]
        
        if not cleaned.startswith('+') and len(cleaned) >= 10:
            if cleaned.startswith('7'):
                cleaned = '+' + cleaned
            else:
                cleaned = '+7' + cleaned
        
        return cleaned
    
    def _get_user_export_data_batch(self, users: List[models.User]) -> Dict[int, Dict[str, Any]]:
        """
        Собирает данные для списка пользователей за один запрос с joinedload.
        Исправляет N+1 проблему.
        """
        if not users:
            return {}
        
        user_ids = [u.id for u in users]
        
        # Получаем все данные одним запросом с подгрузкой связанных таблиц
        users_with_data = self.db.query(models.User).options(
            joinedload(models.User.work),
            joinedload(models.User.education),
            joinedload(models.User.additional_info)
        ).filter(models.User.id.in_(user_ids)).all()
        
        # Создаём словарь для быстрого доступа
        users_map = {u.id: u for u in users_with_data}
        
        result = {}
        
        for user in users:
            user_data = users_map.get(user.id)
            if not user_data:
                continue
            
            # Получаем текущее место работы
            current_work = None
            if user_data.work:
                for w in user_data.work:
                    if w.is_current:
                        current_work = w
                        break
                if not current_work:
                    current_work = user_data.work[0] if user_data.work else None
            
            # Получаем основное образование
            main_education = None
            if user_data.education:
                for e in user_data.education:
                    if e.is_main:
                        main_education = e
                        break
                if not main_education:
                    main_education = user_data.education[0] if user_data.education else None
            
            additional_info = user_data.additional_info
            
            subjects = []
            if current_work and current_work.subjects:
                try:
                    subjects = json.loads(current_work.subjects)
                    if not isinstance(subjects, list):
                        subjects = []
                except (json.JSONDecodeError, TypeError):
                    subjects = []
            
            birth_date_str = ""
            if user.birth_date:
                try:
                    birth_date_str = user.birth_date.strftime("%d.%m.%Y")
                except Exception:
                    birth_date_str = ""
            
            gender_str = ""
            if user.gender == "male":
                gender_str = "Мужской"
            elif user.gender == "female":
                gender_str = "Женский"
            
            snils_value = ""
            if additional_info and additional_info.snils:
                snils_value = self._sanitize_snils(additional_info.snils)
            
            phone_value = user.phone_raw or user.phone or ""
            phone_value = self._sanitize_phone(phone_value)
            
            result[user.id] = {
                "municipality": self._escape_excel_string(user.municipality),
                "organization": self._escape_excel_string(current_work.organization if current_work else ""),
                "last_name": self._escape_excel_string(user.last_name),
                "first_name": self._escape_excel_string(user.first_name),
                "middle_name": self._escape_excel_string(user.middle_name),
                "birth_date": self._escape_excel_string(birth_date_str),
                "gender": self._escape_excel_string(gender_str),
                "snils": self._escape_excel_string(snils_value),
                "position": self._escape_excel_string(current_work.position if current_work else ""),
                "activity_type": self._escape_excel_string(current_work.activity_type if current_work else ""),
                "subjects": self._escape_excel_string("; ".join(subjects) if subjects else ""),
                "education": self._escape_excel_string(main_education.education_level if main_education else ""),
                "document_series": self._escape_excel_string(main_education.document_series if main_education else ""),
                "document_number": self._escape_excel_string(main_education.document_number if main_education else ""),
                "teaching_experience": current_work.teaching_experience_years if current_work and current_work.teaching_experience_years is not None else "",
                "work_experience": current_work.work_experience_years if current_work and current_work.work_experience_years is not None else "",
                "email": self._escape_excel_string(user.email),
                "phone": self._escape_excel_string(phone_value)
            }
        
        return result
    
    def _get_headers(self) -> List[str]:
        return [
            "Муниципалитет",
            "Место работы",
            "Фамилия",
            "Имя",
            "Отчество",
            "Дата рождения получателя",
            "Пол получателя",
            "СНИЛС",
            "Должность",
            "Вид деятельности",
            "Предмет",
            "Образование",
            "Серия диплома",
            "Номер диплома",
            "Педагогический стаж",
            "Стаж в должности",
            "Личный электронный адрес",
            "Номер телефона"
        ]
    
    def export_users_to_excel(self, user_ids: List[int] = None) -> BytesIO:
        try:
            if user_ids:
                users = self.db.query(models.User).filter(models.User.id.in_(user_ids)).all()
            else:
                users = self.db.query(models.User).all()
            
            if not users:
                return self._get_empty_template()
            
            # Получаем данные для всех пользователей одним запросом
            users_data = self._get_user_export_data_batch(users)
            
            wb = self._load_template()
            ws = wb.active
            
            start_row = 5
            
            for row_idx, user in enumerate(users, start=start_row):
                data = users_data.get(user.id, {})
                
                for field, col_idx in self.COLUMN_MAPPING.items():
                    value = data.get(field, "")
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            for row in range(start_row, start_row + len(users)):
                ws.row_dimensions[row].height = 30
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            logger.info(f"Exported {len(users)} users to Excel")
            return output
            
        except Exception as e:
            logger.error(f"Error exporting users to Excel: {str(e)}")
            raise
    
    def export_single_user(self, user_id: int) -> Optional[BytesIO]:
        try:
            user = self.db.query(models.User).filter(models.User.id == user_id).first()
            if not user:
                return None
            
            return self.export_users_to_excel([user_id])
        except Exception as e:
            logger.error(f"Error exporting single user {user_id}: {str(e)}")
            return None
    
    def _load_template(self) -> Any:
        template_path = os.path.join(os.getcwd(), self.TEMPLATE_PATH)
        
        if os.path.exists(template_path):
            return load_workbook(template_path)
        else:
            return self._create_template()
    
    def _create_template(self) -> Any:
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Данные пользователей"
        
        headers = self._get_headers()
        header_row = 4
        
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='0057A4', end_color='0057A4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for col_idx, header in enumerate(headers, start=3):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        column_widths = {
            1: 6, 2: 6, 3: 30, 4: 50, 5: 22, 6: 22, 7: 22,
            8: 22, 9: 16, 10: 20, 11: 30, 12: 25, 13: 35,
            14: 30, 15: 20, 16: 20, 17: 18, 18: 18, 19: 35, 20: 20
        }
        
        for col_idx, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        ws.row_dimensions[header_row].height = 35
        ws.freeze_panes = 'C5'
        
        os.makedirs(os.path.dirname(self.TEMPLATE_PATH), exist_ok=True)
        wb.save(self.TEMPLATE_PATH)
        
        return wb
    
    def _get_empty_template(self) -> BytesIO:
        wb = self._load_template()
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def get_users_list_with_data(self, limit: int = None, offset: int = 0) -> Dict[str, Any]:
        """
        Получает список пользователей с их данными.
        Оптимизировано: все связанные данные загружаются одним запросом.
        Возвращает объект с пагинацией.
        """
        try:
            # Базовый запрос с подгрузкой связанных данных
            query = self.db.query(models.User).options(
                joinedload(models.User.work),
                joinedload(models.User.education),
                joinedload(models.User.additional_info)
            )
            
            total = query.count()
            
            if limit is not None:
                query = query.offset(offset).limit(limit)
            
            users = query.all()
            
            # Получаем данные для всех пользователей одним запросом
            users_data = self._get_user_export_data_batch(users)
            
            result = []
            for user in users:
                data = users_data.get(user.id, {})
                
                full_name = f"{user.last_name or ''} {user.first_name or ''} {user.middle_name or ''}".strip()
                if not full_name:
                    full_name = user.full_name or ""
                
                result.append({
                    "id": user.id,
                    "email": self._escape_excel_string(user.email),
                    "role": user.role.value if user.role else "teacher",
                    "is_blocked": user.is_blocked,
                    "created_at": user.created_at.strftime("%d.%m.%Y %H:%M") if user.created_at else "",
                    "full_name": self._escape_excel_string(full_name),
                    "data": data,
                    "has_complete_profile": user.is_profile_complete()
                })
            
            return {
                "total": total,
                "items": result
            }
            
        except Exception as e:
            logger.error(f"Error getting users list: {str(e)}")
            return {"total": 0, "items": []}


def generate_export_filename(export_type: str = "all", user_id: int = None) -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    if export_type == "single" and user_id:
        return f"user_{user_id}_data_{timestamp}.xlsx"
    else:
        return f"all_users_data_{timestamp}.xlsx"